"""Extract selected RCPLIB solution columns to JSON.

The workbook stores benchmark instance names in column E and summary values in
far-right columns. This script extracts columns AP, BG, and BH from the selected
sheets and writes a flat JSON mapping:

{
  "instance_name": {
    "CPM": 99,
    "LB-lit": 105,
    "UB-lit": 105
  }
}

By default the second-level keys come from the header row of AP/BG/BH. Use
``--key-style letter`` if the JSON should use ``AP``, ``BG``, and ``BH`` as
keys instead.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_WORKBOOK = SCRIPT_DIR / "RCPLIB_Solution.xlsx"
DEFAULT_OUTPUT = SCRIPT_DIR / "rcplib_solution_results.json"
DEFAULT_SHEETS = ["PSPLIB", "LPP", "LPSP", "RG300"]
SHEET_ALIASES = {
    "GR300": "RG300",
}
KEY_COLUMN = "E"
VALUE_COLUMNS = ["AP", "BG", "BH"]


def normalize_excel_value(value: Any) -> Any:
    """Convert openpyxl scalar values to stable JSON-friendly values."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def resolve_sheet_name(requested_sheet: str, workbook_sheets: list[str]) -> str:
    """Return the workbook sheet name, allowing known aliases."""
    if requested_sheet in workbook_sheets:
        return requested_sheet

    alias = SHEET_ALIASES.get(requested_sheet)
    if alias and alias in workbook_sheets:
        return alias

    available = ", ".join(workbook_sheets)
    raise ValueError(
        f"Sheet {requested_sheet!r} was not found. Available sheets: {available}"
    )


def extract_solution_columns(
    workbook_path: str | Path,
    sheets: list[str] | None = None,
    key_style: str = "header",
) -> dict[str, dict[str, Any]]:
    """Extract E/AP/BG/BH values from the requested workbook sheets."""
    if key_style not in {"header", "letter"}:
        raise ValueError("key_style must be 'header' or 'letter'.")

    requested_sheets = sheets or DEFAULT_SHEETS
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    workbook_sheets = wb.sheetnames

    key_col_idx = column_index_from_string(KEY_COLUMN)
    value_col_indices = [column_index_from_string(col) for col in VALUE_COLUMNS]
    min_col = min([key_col_idx, *value_col_indices])
    max_col = max([key_col_idx, *value_col_indices])

    extracted: dict[str, dict[str, Any]] = {}
    source_for_key: dict[str, str] = {}

    for requested_sheet in requested_sheets:
        sheet_name = resolve_sheet_name(requested_sheet, workbook_sheets)
        ws = wb[sheet_name]

        rows = ws.iter_rows(
            min_row=1,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        )

        try:
            header_row = next(rows)
        except StopIteration:
            continue

        def value_at(column_idx: int, row: tuple[Any, ...]) -> Any:
            return row[column_idx - min_col]

        if key_style == "header":
            value_keys = [
                str(normalize_excel_value(value_at(col_idx, header_row)))
                for col_idx in value_col_indices
            ]
        else:
            value_keys = VALUE_COLUMNS

        for row_number, row in enumerate(rows, start=2):
            raw_key = value_at(key_col_idx, row)
            if raw_key is None or raw_key == "":
                continue

            instance_key = str(normalize_excel_value(raw_key))
            if instance_key in extracted:
                previous = source_for_key[instance_key]
                current = f"{sheet_name}!{KEY_COLUMN}{row_number}"
                raise ValueError(
                    f"Duplicate first-level key {instance_key!r}: "
                    f"{previous} and {current}."
                )

            extracted[instance_key] = {
                value_key: normalize_excel_value(value_at(col_idx, row))
                for value_key, col_idx in zip(
                    value_keys,
                    value_col_indices,
                    strict=True,
                )
            }
            source_for_key[instance_key] = f"{sheet_name}!{KEY_COLUMN}{row_number}"

    return extracted


def write_solution_json(
    workbook_path: str | Path,
    output_path: str | Path,
    sheets: list[str] | None = None,
    key_style: str = "header",
) -> dict[str, dict[str, Any]]:
    """Extract workbook values and write them to ``output_path`` as JSON."""
    extracted = extract_solution_columns(
        workbook_path=workbook_path,
        sheets=sheets,
        key_style=key_style,
    )
    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(extracted, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    return extracted


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract E/AP/BG/BH columns from RCPLIB_Solution.xlsx."
    )
    parser.add_argument(
        "--input",
        type=Path,
        default=DEFAULT_WORKBOOK,
        help=f"Workbook path (default: {DEFAULT_WORKBOOK}).",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Output JSON path (default: {DEFAULT_OUTPUT}).",
    )
    parser.add_argument(
        "--sheets",
        nargs="+",
        default=DEFAULT_SHEETS,
        help=(
            "Sheet names to extract. GR300 is still accepted as an alias for "
            "the workbook's RG300 sheet. Default: PSPLIB LPP LPSP RG300."
        ),
    )
    parser.add_argument(
        "--key-style",
        choices=["header", "letter"],
        default="header",
        help=(
            "Use AP/BG/BH header values as second-level keys, or use the "
            "literal Excel column letters (default: header)."
        ),
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    extracted = write_solution_json(
        workbook_path=args.input,
        output_path=args.output,
        sheets=args.sheets,
        key_style=args.key_style,
    )
    print(f"Wrote {len(extracted)} entries to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
